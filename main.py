import dimensions
import metrics
import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            </style>
            """

st.markdown("""
<style>
.big-font {
    font-size:50px !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""<h1>Keyword Cannibalization Tool</h1> <p>Find pages that are competing with each other for the same 
keyword using your first-party GSC Data or SEMrush Keyword data. The data must be multi-dimensional and contain 
query (keyword) and page data.</p> <b>Directions: </b> <ul> <li>Upload query + page data (.csv) from the GSC API or SEMrush Keyword Export.</li> 
<li>Optional: Set a threshold to select the top n% of queries (default 80%).</li> </ul>

</ul><br><br>
""", unsafe_allow_html=True)

perc_slider = st.slider('Set Threshold (ex: 80 = Selecting the top 80% of queries by metric)', 0, 100, value=80,
                        step=10, key='perc_slider')
gsc_data_file = st.file_uploader('Upload Data', type='csv', key='key')

summary = []
def check_source(df):
    # check data source
    if 'query' in df.columns.str.lower() and 'clicks' in df.columns.str.lower():
        if 'page' in df.columns.str.lower():
            return dimensions.gsc_dimensions, metrics.gsc_metrics
        else:
            st.error('Data must contain a "page" column.')
            st.stop()
    elif 'keyword' in df.columns.str.lower() and 'url' in df.columns.str.lower():
        return dimensions.semrush_dimensions, metrics.semrush_metrics
    else:
        st.error('Data file is invalid. Please upload a valid data file.')
        st.stop()


def is_ascii(string):
    # remove rows containing non-ascii characters
    try:
        string.encode('ascii')
    except UnicodeEncodeError:
        return False
    else:
        return True


def process_data(df, metric, perc_cumsum, dimension):
    dimension = [d.lower() for d in dimension]
    groupby_column_name = dimension[0]
    metric_name = metric.lower()

    # format column names for final DataFrame
    if 'query' in dimension:
        df_cols = [groupby_column_name, 'page', metric_name, 'ctr', 'position', metric_name + '_percent_all_query',
                   'query_percentile_' + metric_name]
    if 'keyword' in dimension:
        df_cols = [groupby_column_name, 'url', metric_name, 'cpc', 'position', metric_name + '_percent_all_keyword']

    # Group by the specified metric
    grouped_df = df.groupby(metric_name, as_index=False).apply(
        lambda group: group.sort_values(metric_name, ascending=False))

    # Calculate sum per group and percentage per group
    sum_per_group = grouped_df.groupby(groupby_column_name)[metric_name].sum().sort_values(ascending=False)
    percent_per_group = grouped_df.groupby(groupby_column_name)[metric_name].sum() / grouped_df[metric_name].sum()

    # Calculate cumulative sum and get top 90% of groups
    cumsum_percent = percent_per_group.sort_values(ascending=False).cumsum()
    top_n_percent = percent_per_group[cumsum_percent <= float(perc_cumsum)]

    # Merge sum_per_group and top_90_percent back to the original DataFrame
    df = pd.merge(grouped_df, sum_per_group, left_on=groupby_column_name, right_index=True, suffixes=('', '_sum'))
    df = pd.merge(df, top_n_percent, left_on=groupby_column_name, right_index=True,
                  suffixes=('', f'_percent_all_{dimension[0]}'))

    # Sort DataFrame
    df.sort_values([metric_name + '_sum', metric_name], ascending=[False, False], inplace=True)

    # Group by query and calculate the % total of each row within each group
    df[f'{dimension[0]}_percentile_' + metric_name] = df[metric_name] / df[metric_name + '_sum']
    df.sort_values([metric_name + '_sum', metric_name], ascending=[False, False], inplace=True)

    # Keep rows where query_percentile is greater than or equal to 0.09
    df = df[df[f'{dimension[0]}_percentile_' + metric_name] >= 0.10]

    # Drop duplicates rows of column 1 and column 2
    df.drop_duplicates(subset=[dimension[0], dimension[1]], inplace=True)

    # Keep only duplicate rows in the query/keyword column
    df = df[df.duplicated(subset=[dimension[0]], keep=False)]

    # Sort by metric_name + sum, then by metric_name, then by query
    df.sort_values([metric_name + '_sum', dimension[0], metric_name, 'position'], ascending=[False, True, False, True],
                   inplace=True)

    unique_values = df[dimension[0]].nunique()
    summary.append(f'{unique_values} unique keywords competing by {metric_name}.')

    # Select specific columns
    df = df[df_cols]
    return df


def process_merge(dfs, dimension):
    merged_df = pd.merge(dfs[0], dfs[1], on=[dimension[0], dimension[1]], how='inner')
    merged_df = merged_df[merged_df.duplicated(subset=[dimension[0]], keep=False)]

    # remove duplicate columns
    merged_df = merged_df.loc[:, ~merged_df.columns.str.endswith('_y')]
    merged_df = merged_df.rename(columns=lambda x: x.replace('_x', ''))
    return merged_df


def format_excel(xlsx_file):
    wb = load_workbook(filename=xlsx_file)

    # Create a new objects for "Good" rows
    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    good_font = Font(color="006100")

    for ws in wb.worksheets:
        # Set zoom to 130%
        ws.sheet_view.zoomScale = 130

        # Keep track of queries already seen
        seen_queries = set()

        # Iterate through rows to find and highlight the first instance of each query
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row, values_only=False):
            query = row[0].value
            if query not in seen_queries:
                seen_queries.add(query)
                for cell in row:
                    cell.fill = good_fill
                    cell.font = good_font

        # Rename sheet
        ws.title = f'Competing by {ws.title}'

    # Save the modified workbook
    wb.save(f'cannibalization_data_threshold_{perc_slider}.xlsx')
    return wb


if __name__ == '__main__':
    if gsc_data_file is not None:
        data = pd.read_csv(gsc_data_file)

        # force column names lowercase
        data.columns = data.columns.str.lower()

        # check data source
        dimensions, metrics = check_source(data)
        data = data[data[dimensions[0]].apply(lambda x: is_ascii(str(x)))]
        perc_cumsum = perc_slider / 100

        # remove rows with 0 clicks or traffic
        data = data[data[metrics[0]] > 0]
        dfs = []
        wb = Workbook()
        for metric in metrics[:2]:
            df_processed = process_data(data, metric, perc_cumsum, dimensions)
            dfs.append(df_processed)

        # merge dfs
        dfs.append(process_merge(dfs, dimensions))
        unique_vals = dfs[2][dimensions[0]].nunique()
        summary.append(f'{unique_vals} unique keywords competing by {metrics[2]}.')
        try:
            for sheet_name, df in zip(metrics, dfs):
                sheet = wb.create_sheet(title=sheet_name)
                # write dataframe to the sheet

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
        except IndexError:
            st.error('Threshold is too low. Please increase the threshold.')
        wb.remove(wb['Sheet'])
        wb.save(f'cannibalization_data_threshold_{perc_slider}.xlsx')
        # format excel
        wb = format_excel(f'cannibalization_data_threshold_{perc_slider}.xlsx')
        wb.save(f'cannibalization_data_threshold_{perc_slider}.xlsx')
        # display summary
        st.markdown(f'### Summary, threshold set to {perc_slider}%')
        for s in summary:
            st.markdown(f'{s}')
        st.markdown(f'')
        with open(f"cannibalization_data_threshold_{perc_slider}.xlsx", "rb") as file:
            st.download_button(label=f'Download Cannibalization Analysis',
                               data=file,
                               file_name=f'cannibalization_data_threshold_{perc_slider}.xlsx',
                               mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                               key='download_button')

st.write('---')
st.markdown(hide_streamlit_style, unsafe_allow_html=True)
st.write(
    'Author: [Tyler Gargula](https://tylergargula.dev) | Technical SEO & Software Developer | [Buy Me a Coffee](https://venmo.com/u/Tyler-Gargula)️☕️')
